import numpy as np
from sklearn import preprocessing
import numpy as np
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

wb = load_workbook(r"C:\Users\alqas\Research\SHEtable2Vdc.xlsx")
ws = wb.active

x_vdc1=[]
y_vdc2=[]
z_t1=[[],[],[],[],[],[]]
#collecting the data from the columns we need
for col in ws.iter_cols(min_row=2, max_col=2 ,min_col=2, values_only=True):
    for row in col:
        x_vdc1.append(row)
for col in ws.iter_cols(min_row=2, max_col=3 ,min_col=3, values_only=True):
    for row in col:
        y_vdc2.append(row)
for g in range(1,7):
    for col in ws.iter_cols(min_row=2, max_col=g+3 ,min_col=g+3, values_only=True):
        for row in col:
            z_t1[g-1].append(row)

#setting up the arrays to be able to handle matrix math, using np
x_vdc1 = np.array(x_vdc1)
y_vdc2 = np.array(y_vdc2)
z_t1 = np.array(z_t1)
x_vdc1.reshape(-1,1)
y_vdc2.reshape(-1,1)
output = np.column_stack((z_t1[0]))
if len(output)==1:
    output = [[el] for el in z_t1[0]]

input = np.column_stack((x_vdc1,y_vdc2))
#scalling works best for gradient decent unscaling will be done at the end
x = preprocessing.MinMaxScaler()
y = preprocessing.MinMaxScaler()

scaled_input = x.fit_transform(input)
scaled_output = y.fit_transform(output)

#this function will use the gradient decent equations to do the gradient decent
#still working on making the function work for multiple outputs at the moment
def gradient_decent(X,y_true,epoch,learning_rate = .01):
    number_of_features = X.shape[1]
    w = np.ones(shape = number_of_features)
    b = 0
    total_sample = X.shape[0]

    cost_list = []
    epoch_list = []
    for i in range(epoch):
        y_predict = b + np.dot(w,scaled_input.T)

        w_grad = -(2/total_sample)*(X.T.dot(y_true-y_predict))
        b_grad = -(2/total_sample)*np.sum(y_true-y_predict)

        w = w- learning_rate*w_grad
        b = b-learning_rate*b_grad
        cost = np.mean(np.square(y_true-y_predict))
        
        if i%10 == 0:
            cost_list.append(cost)
            epoch_list.append(i)
    return w,b,cost,cost_list, epoch_list
#using the gradient decent function just created
w , b , cost,cost_list , epoch_list =gradient_decent(scaled_input, scaled_output.reshape(scaled_output.shape[0],),500)

#ploting the gradient decent function, to show its progress(not needed)
plt.plot(epoch_list, cost_list)
plt.show()

#ceating a function that predicts and reverse scales the values
def predict(v_1, v_2, w,b):
    scaled_input = x.transform([[v_1, v_2]])[0]
    scaled_output_1 = w[0]*scaled_input[0]+w[1]*scaled_input[1]+b
    return y.inverse_transform([[scaled_output_1]])
    
#predicting the values is relativley accurate
print(predict(112,120,w,b))
