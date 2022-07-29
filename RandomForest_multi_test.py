from tkinter import N
import numpy as np
import numpy as np
from openpyxl import load_workbook
import numpy as np
import timeit
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
from matplotlib import cm
import pandas as pd
from sklearn.metrics import mean_squared_error
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import cross_val_score

N_Estimators = 500
Max_Leaf_Nodes = 16
N_Jobes= -1

wb = load_workbook(r"C:\Users\alqas\Research\SHEtable2Vdc.xlsx")
ws = wb.active

x_vdc1=list()
y_vdc2=list()
z_t1=[[],[],[],[],[],[]]
#collecting the data from the columns we need
for col in ws.iter_cols(min_row=2, max_col=2 ,min_col=2, values_only=True):
    for row in col:
        x_vdc1.append(row)
for col in ws.iter_cols(min_row=2, max_col=3 ,min_col=3, values_only=True):
    for row in col:
        y_vdc2.append(row)
for g in range(1,7):
    for col in ws.iter_cols(min_row=2, max_col=g+3 ,min_col=g+3, values_only=True):
        for row in col:
            z_t1[g-1].append(row)

# resizing the data to fit the model
df = pd.DataFrame({"x_1":np.array(x_vdc1).reshape(496,), "x_2":np.array(y_vdc2).reshape(496,), "y_1":np.array(z_t1[0]).reshape(496,), "y_2":np.array(z_t1[1]).reshape(496,), "y_3":np.array(z_t1[2]).reshape(496,), "y_4":np.array(z_t1[3]).reshape(496,), "y_5":np.array(z_t1[4]).reshape(496,), "y_6":np.array(z_t1[5]).reshape(496,)}, index=range(0,496))

#sepperating inputs and outputs
X, y = df[["x_1", "x_2"]], df[["y_1","y_2","y_3","y_4","y_5","y_6"]]

#splitting test and training data
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=48)

#fitting RandomForest regression
RandomForest_Reg = RandomForestRegressor(n_estimators=N_Estimators, max_leaf_nodes= Max_Leaf_Nodes,n_jobs= N_Jobes )
RandomForest_Reg.fit(X_train,y_train)

#predicting values for test set
start=timeit.default_timer()  
RandomForest_predicted = RandomForest_Reg.predict(X_test)

#finding Mean squared error
RandomForest_rmse = np.sqrt(mean_squared_error(y_test, RandomForest_predicted))
stop = timeit.default_timer()

#reformating input data for graphing
x_test = X_test["x_1"]
Y_test = X_test["x_2"]
differance = RandomForest_predicted - y_test

#graphing expected vs predicted data
# fig = plt.figure()
# ax_real = fig.add_subplot(1,3,1,projection='3d')
# ax_prediction = fig.add_subplot(1,3,2, projection = '3d')
# ax_differance = fig.add_subplot(1,3,3, projection = '3d')
# surf_real = ax_real.plot_trisurf(x_test,Y_test,y_test, cmap=cm.coolwarm,linewidth=0, antialiased=False)
# surf_prediction= ax_prediction.plot_trisurf(x_test,Y_test, RandomForest_predicted, cmap=cm.coolwarm,linewidth=0, antialiased=False)
# surf_differance = ax_differance.plot_trisurf(x_test,Y_test, differance, cmap=cm.coolwarm,linewidth=0, antialiased=False)
# cbaxes = fig.add_axes([0.1, 0.3, 0.01, 0.4]) 
# fig.colorbar(surf_real, shrink=0.5, aspect=5, cax = cbaxes )
# ax_real.set_xlabel('Vdc1, volts')
# ax_real.set_ylabel('Vdc2, volts')
# ax_real.set_zlabel('Angle, degrees')
# ax_prediction.set_xlabel('Vdc1, volts')
# ax_prediction.set_ylabel('Vdc2, volts')
# ax_prediction.set_zlabel('Angle, degrees')
# ax_differance.set_xlabel('Vdc1, volts')
# ax_differance.set_ylabel('Vdc2, volts')
# ax_differance.set_zlabel('Angle, degrees')

#cross validation 
scores = cross_val_score(RandomForest_Reg,X,y,scoring="neg_mean_squared_error",cv=10)
RandomForest_CV = np.sqrt(-scores) 


## EXECUTABLE ITEMS

#predicted values
print("prediction values",RandomForest_predicted)
# RMSE of cross validation
print("RMSE for set",RandomForest_rmse)
# time prediction/execution
print('Time for prediction: ', stop - start)
# Cross validation information
print("the value of cross validation value",RandomForest_CV.mean())
#show graph
plt.show()