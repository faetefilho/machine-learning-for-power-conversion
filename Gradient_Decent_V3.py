import numpy as np
from sklearn import preprocessing
import numpy as np
from openpyxl import load_workbook
import numpy as np
import timeit
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
from matplotlib import cm, projections
from matplotlib.ticker import LinearLocator, FormatStrFormatter, MaxNLocator

wb = load_workbook(r"C:\Users\alqas\Research\SHEtable2Vdc.xlsx")
ws = wb.active


x_vdc1=list()
y_vdc2=list()
z_t1=[[],[],[],[],[],[]]
#collecting the data from the columns we need
for col in ws.iter_cols(min_row=2, max_col=2 ,min_col=2, values_only=True):
    for row in col:
        x_vdc1.append(row)
for col in ws.iter_cols(min_row=2, max_col=3 ,min_col=3, values_only=True):
    for row in col:
        y_vdc2.append(row)
for g in range(1,7):
    for col in ws.iter_cols(min_row=2, max_col=g+3 ,min_col=g+3, values_only=True):
        for row in col:
            z_t1[g-1].append(row)

#setting up the arrays to be able to handle matrix math, using np
train_set_x, test_set_x = train_test_split(x_vdc1,test_size=.2,random_state=48)
train_set_y, test_set_y = train_test_split(y_vdc2,test_size=.2,random_state=48)
train_set_z1, test_set_z1 = train_test_split(z_t1[0],test_size=.2,random_state=48)
train_set_z2, test_set_z2 = train_test_split(z_t1[1],test_size=.2,random_state=48)
train_set_z3, test_set_z3 = train_test_split(z_t1[2],test_size=.2,random_state=48)
train_set_z4, test_set_z4 = train_test_split(z_t1[3],test_size=.2,random_state=48)
train_set_z5, test_set_z5 = train_test_split(z_t1[4],test_size=.2,random_state=48)
train_set_z6, test_set_z6 = train_test_split(z_t1[5],test_size=.2,random_state=48)
train_set_z = [train_set_z1,train_set_z2,train_set_z3,train_set_z4,train_set_z5,train_set_z6]
test_set_z = [test_set_z1, test_set_z2,test_set_z3, test_set_z4,test_set_z5, test_set_z6]

z_variable_selection = 0

train_set_x = np.array(train_set_x)
train_set_y = np.array(train_set_y)
train_set_z = np.array(train_set_z)
train_set_x.reshape(-1,1)
train_set_y.reshape(-1,1)
output = np.column_stack((train_set_z[z_variable_selection]))
if len(output)==1:
    output = [[el] for el in train_set_z[z_variable_selection]]

input = np.column_stack((train_set_x,train_set_y))
#scalling works best for gradient decent unscaling will be done at the end
x = preprocessing.MinMaxScaler()
y = preprocessing.MinMaxScaler()

scaled_input = x.fit_transform(input)
scaled_output = y.fit_transform(output)

#this function will use the gradient decent equations to do the gradient decent
#still working on making the function work for multiple outputs at the moment
def gradient_decent(X,y_true,epoch,learning_rate = .01):
    number_of_features = X.shape[1]
    w = np.ones(shape = number_of_features)
    b = 0
    total_sample = X.shape[0]

    cost_list = []
    epoch_list = []
    for i in range(epoch):
        y_predict = b + np.dot(w,scaled_input.T)

        w_grad = -(2/total_sample)*(X.T.dot(y_true-y_predict))
        b_grad = -(2/total_sample)*np.sum(y_true-y_predict)

        w = w- learning_rate*w_grad
        b = b-learning_rate*b_grad
        cost = np.mean(np.square(y_true-y_predict))
        
        if i%10 == 0:
            cost_list.append(cost)
            epoch_list.append(i)
    return w,b,cost,cost_list, epoch_list
#using the gradient decent function just created
w , b , cost,cost_list , epoch_list =gradient_decent(scaled_input, scaled_output.reshape(scaled_output.shape[0],),500)

#ploting the gradient decent function, to show its progress(not needed)
# plt.plot(epoch_list, cost_list)
# plt.show()

#ceating a function that predicts and reverse scales the values
def predict(v_1, v_2, w,b):
    scaled_input = x.transform([[v_1, v_2]])[0]
    scaled_output_1 = w[0]*scaled_input[0]+w[1]*scaled_input[1]+b
    return y.inverse_transform([[scaled_output_1]])

#predicting the values of all test set
start=timeit.default_timer()
prediction_data = list()
for test_value in range(len(test_set_x)):
    prediction_data.append(float(predict(int(test_set_x[test_value]),int(test_set_y[test_value]),w,b)))

#end of timer for prediction 
stop = timeit.default_timer()
print('Time: ', stop - start)


# mean squared error calculation
# test set
summation = 0
for test_value in range(len(test_set_x)):
    summation = summation + (test_set_z[z_variable_selection][test_value]-prediction_data[test_value])**2
MSE = 1/len(test_set_x)*summation
print(MSE)

fig = plt.figure()
ax_real = fig.add_subplot(1,2,1,projection='3d')
ax_prediction = fig.add_subplot(1,2,2, projection = '3d')
surf_real = ax_real.plot_trisurf(test_set_x,test_set_y,test_set_z[z_variable_selection], cmap=cm.coolwarm,linewidth=0, antialiased=False)
surf_prediction= ax_prediction.plot_trisurf(test_set_x, test_set_y, prediction_data, cmap=cm.coolwarm,linewidth=0, antialiased=False)
cbaxes = fig.add_axes([0.1, 0.3, 0.01, 0.4]) 
fig.colorbar(surf_real, shrink=0.5, aspect=5, cax = cbaxes )
ax_real.set_xlabel('Vdc1, volts')
ax_real.set_ylabel('Vdc2, volts')
ax_real.set_zlabel('Angle, degrees')
ax_prediction.set_xlabel('Vdc1, volts')
ax_prediction.set_ylabel('Vdc2, volts')
ax_prediction.set_zlabel('Angle, degrees')
plt.show()