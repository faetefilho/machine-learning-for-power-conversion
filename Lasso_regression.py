import numpy as np
import numpy as np
from openpyxl import load_workbook
import numpy as np
import timeit
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
from matplotlib import cm
import pandas as pd
from sklearn.metrics import mean_squared_error
from sklearn.linear_model import Lasso

wb = load_workbook(r"C:\Users\alqas\Research\SHEtable2Vdc.xlsx")
ws = wb.active


x_vdc1=list()
y_vdc2=list()
z_t1=[[],[],[],[],[],[]]
#collecting the data from the columns we need
for col in ws.iter_cols(min_row=2, max_col=2 ,min_col=2, values_only=True):
    for row in col:
        x_vdc1.append(row)
for col in ws.iter_cols(min_row=2, max_col=3 ,min_col=3, values_only=True):
    for row in col:
        y_vdc2.append(row)
for g in range(1,7):
    for col in ws.iter_cols(min_row=2, max_col=g+3 ,min_col=g+3, values_only=True):
        for row in col:
            z_t1[g-1].append(row)

# resizing the data to fit the model
z_variable_selection = 0
df = pd.DataFrame({"x_1":np.array(x_vdc1).reshape(496,), "x_2":np.array(y_vdc2).reshape(496,), "y":np.array(z_t1[z_variable_selection]).reshape(496,)}, index=range(0,496))

#sepperating inputs and outputs
X, y = df[["x_1", "x_2"]], df["y"]

#splitting test and training data
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=48)

#fitting Lasso regression
Lasso_Reg = Lasso(alpha=.1)
Lasso_Reg.fit(X_train,y_train)

#predicting values for test set
start=timeit.default_timer()  
Lasso_predicted = Lasso_Reg.predict(X_test)

#finding Mean squared error
Lasso_rmse = np.sqrt(mean_squared_error(y_test, Lasso_predicted))

stop = timeit.default_timer()
print('Time: ', stop - start)

#reformating input data for graphing
x_test = X_test["x_1"]
Y_test = X_test["x_2"]

print(Lasso_predicted)
print(Lasso_rmse)

#graphing expected vs predicted data
fig = plt.figure()
ax_real = fig.add_subplot(1,2,1,projection='3d')
ax_prediction = fig.add_subplot(1,2,2, projection = '3d')
surf_real = ax_real.plot_trisurf(x_test,Y_test,y_test, cmap=cm.coolwarm,linewidth=0, antialiased=False)
surf_prediction= ax_prediction.plot_trisurf(x_test,Y_test, Lasso_predicted, cmap=cm.coolwarm,linewidth=0, antialiased=False)
cbaxes = fig.add_axes([0.1, 0.3, 0.01, 0.4]) 
fig.colorbar(surf_real, shrink=0.5, aspect=5, cax = cbaxes )
ax_real.set_xlabel('Vdc1, volts')
ax_real.set_ylabel('Vdc2, volts')
ax_real.set_zlabel('Angle, degrees')
ax_prediction.set_xlabel('Vdc1, volts')
ax_prediction.set_ylabel('Vdc2, volts')
ax_prediction.set_zlabel('Angle, degrees')
plt.show()