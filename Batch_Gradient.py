from turtle import pd
from sklearn.datasets import make_regression
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
import pandas as pd
from sklearn import linear_model
import numpy as np
from sklearn.linear_model import SGDRegressor



wb = load_workbook(r"C:\Users\alqas\Research\SHEtable2Vdc.xlsx")
ws = wb.active

x_vdc1=[]
y_vdc2=[]
z_t1=[[],[],[],[],[],[]]
#collecting the data from the columns we need
for col in ws.iter_cols(min_row=2, max_col=2 ,min_col=2, values_only=True):
    for row in col:
        x_vdc1.append(row)
for col in ws.iter_cols(min_row=2, max_col=3 ,min_col=3, values_only=True):
    for row in col:
        y_vdc2.append(row)
for g in range(1,7):
    for col in ws.iter_cols(min_row=2, max_col=g+3 ,min_col=g+3, values_only=True):
        for row in col:
            z_t1[g-1].append(row)
x_vdc1 = np.array(x_vdc1)
y_vdc2 = np.array(y_vdc2)

Data = {'x': x_vdc1,
        'y': y_vdc2,
        'z_1': z_t1[0],
        'z_2': z_t1[1],
        'z_3': z_t1[2],
        'z_4': z_t1[3],
        'z_5': z_t1[4],
        'z_6': z_t1[5],
        }
df = pd.DataFrame(Data,columns=['x','y','z_1','z_2','z_3','z_4','z_5','z_6'])
input=[x_vdc1,y_vdc2]
output = [z_t1[0],z_t1[1],z_t1[2],z_t1[3]]
input = np.array(input, np.float16)
output = np.array(output, np.float16)
input.resize(1,1)
output.resize(1,1)
def generateXvector(X):
    """ Taking the original independent variables matrix and add a row of 1 which corresponds to x_0
        Parameters:
          X:  independent variables matrix
        Return value: the matrix that contains all the values in the dataset, not include the outcomes variables. 
    """
    vectorX = np.c_[np.ones((len(X), 1)), X]
    return vectorX
def theta_init(X):
    """ Generate an initial value of vector θ from the original independent variables matrix
         Parameters:
          X:  independent variables matrix
        Return value: a vector of theta filled with initial guess
    """
    theta = np.random.randn(len(X[0])+1, 1)
    return theta
def Multivariable_Linear_Regression(X,y,learningrate, iterations):
    """ Find the multivarite regression model for the data set
         Parameters:
          X: independent variables matrix
          y: dependent variables matrix
          learningrate: learningrate of Gradient Descent
          iterations: the number of iterations
        Return value: the final theta vector and the plot of cost function
    """
    y_new = np.reshape(y, (len(y), 1))   
    cost_lst = []
    vectorX = generateXvector(X)
    theta = theta_init(X)
    m = len(X)
    for i in range(iterations):
        gradients = 2/m * vectorX.T.dot(vectorX.dot(theta) - y_new)
        theta = theta - learningrate * gradients
        y_pred = vectorX.dot(theta)
        cost_value = 1/(2*len(y))*((y_pred - y)**2) 
        #Calculate the loss for each training instance
        total = 0
        for i in range(len(y)):
            total += cost_value[i][0] 
            #Calculate the cost function for each iteration
        cost_lst.append(total)
    return theta

print(Multivariable_Linear_Regression(input,output,.000001,1000))

