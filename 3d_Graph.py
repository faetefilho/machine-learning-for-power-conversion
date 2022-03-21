from matplotlib import cm
from matplotlib.ticker import LinearLocator, FormatStrFormatter
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

#creating new listes in order to append data onto them
#in this case the data will come as a tuple
#the data can just be a list, if you extract the value from each cell in 
#the coloumn and apend onto the list, but for this application it is not needed.
x_vdc1=[]
y_vdc2=[]
z_t1=[]

#opening the excel sheet
wb = load_workbook(r"C:\Users\alqas\Research\SHEtable2Vdc.xlsx")
ws = wb.active

#collecting the data from the columns we need
for col in ws.iter_cols(min_row=2, max_col=2 ,min_col=2, values_only=True):
    x_vdc1.append(col)
for col in ws.iter_cols(min_row=2, max_col=3 ,min_col=3, values_only=True):
    y_vdc2.append(col)
for col in ws.iter_cols(min_row=2, max_col=4 ,min_col=4, values_only=True):
    z_t1.append(col)

#defining the graph we will create
fig = plt.figure()
ax = fig.add_subplot(projection='3d')

#ploting the data sets(note since the data is a tuple inside 
#a list we must index the first list vaiable when extracting the data)
surf = ax.plot_trisurf(x_vdc1[0],y_vdc2[0],z_t1[0], cmap=cm.coolwarm,linewidth=0, antialiased=False)

#setting some more deffinition to the Z axis, as well as some cosmetics
ax.zaxis.set_major_locator(LinearLocator(10))
ax.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))
fig.colorbar(surf, shrink=0.5, aspect=5)

#labaling
# remove color bar
plt.title('Switching Angles (\u03B8\N{SUBSCRIPT ONE})')
ax.set_xlabel('Vdc\N{SUBSCRIPT ONE} (V)')
ax.set_ylabel('Vdc\N{SUBSCRIPT TWO} (V)')
ax.set_zlabel('Angle (Degrees)')
plt.show()

# add a plot with columns L M N O P on the same plot and with legends
# add separate plot for column Q

