from turtle import position
from matplotlib import cm, figure
from matplotlib.ticker import LinearLocator, FormatStrFormatter, MaxNLocator
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MaxNLocator

#creating new listes in order to append data onto them
#in this case the data will come as a tuple
#the data can just be a list, if you extract the value from each cell in 
#the coloumn and apend onto the list, but for this application it is not needed.
x_vdc1=[]
y_vdc2=[]
z_t1=[[],[],[],[],[],[]]


#opening the excel sheet
wb = load_workbook(r"C:\Users\alqas\Research\SHEtable2Vdc.xlsx")
ws = wb.active

#collecting the data from the columns we need
for col in ws.iter_cols(min_row=2, max_col=2 ,min_col=2, values_only=True):
    for row in col:
        x_vdc1.append(row)
for col in ws.iter_cols(min_row=2, max_col=3 ,min_col=3, values_only=True):
    for row in col:
        y_vdc2.append(row)
for g in range(1,7):
    for col in ws.iter_cols(min_row=2, max_col=g+3 ,min_col=g+3, values_only=True):
        for row in col:
            z_t1[g-1].append(row)


#defining the graph we will create
fig = plt.figure()

# I felt a class is probably the best way to carry this opperation out on multiple graphs 
# This way even fromating is just a matter of changing the class
class figures:
    def __init__(self,position,zvalue):
        self.position = position
        self.zvalue = zvalue
    def graph(self):
        a=self.zvalue
        b=self.position
        self = fig.add_subplot(2,3,self.position,projection='3d')
        surf = self.plot_trisurf(x_vdc1,y_vdc2,z_t1[a], cmap=cm.coolwarm,linewidth=0, antialiased=False)
        
        
        #cosmetic changes
        self.zaxis.set_major_locator(MaxNLocator(integer=True))
        self.locator_params(axis='z', nbins=4)
        plt.title('Switching Angles (\u03B8\N{SUBSCRIPT ONE}) t'+ str(b),y=1.0 , pad=4)
        self.set_xlabel('Vdc\N{SUBSCRIPT ONE} (V)')
        self.set_ylabel('Vdc\N{SUBSCRIPT TWO} (V)')
        self.set_zlabel('Angle ($^\circ$)')
        fig.colorbar(surf, shrink=0.5, aspect=20, location="left", pad=-.01 )

#deffining each of the functions using the newly created class
p1 = figures(1,0)
p1.graph()
p2 = figures(2,1)
p2.graph()
p3 = figures(3,2)
p3.graph()
p4 = figures(4,3)
p4.graph()
p5 = figures(5,4)
p5.graph()
p6 = figures(6,5)
p6.graph()

plt.show()

