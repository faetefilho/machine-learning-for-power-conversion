from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
import pandas as pd
from sklearn import linear_model
import timeit
import timeit
from sklearn.model_selection import train_test_split
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import cm, projections
from matplotlib.ticker import LinearLocator, FormatStrFormatter, MaxNLocator


wb = load_workbook(r"C:\Users\alqas\Research\SHEtable2Vdc.xlsx")
ws = wb.active

x_vdc1=[]
y_vdc2=[]
z_t1=[[],[],[],[],[],[]]
#collecting the data from the columns we need
for col in ws.iter_cols(min_row=2, max_col=2 ,min_col=2, values_only=True):
    for row in col:
        x_vdc1.append(row)
for col in ws.iter_cols(min_row=2, max_col=3 ,min_col=3, values_only=True):
    for row in col:
        y_vdc2.append(row)
for g in range(1,7):
    for col in ws.iter_cols(min_row=2, max_col=g+3 ,min_col=g+3, values_only=True):
        for row in col:
            z_t1[g-1].append(row)

train_set_x, test_set_x = train_test_split(x_vdc1,test_size=.2,random_state=48)
train_set_y, test_set_y = train_test_split(y_vdc2,test_size=.2,random_state=48)
train_set_z1, test_set_z1 = train_test_split(z_t1[0],test_size=.2,random_state=48)
train_set_z2, test_set_z2 = train_test_split(z_t1[1],test_size=.2,random_state=48)
train_set_z3, test_set_z3 = train_test_split(z_t1[2],test_size=.2,random_state=48)
train_set_z4, test_set_z4 = train_test_split(z_t1[3],test_size=.2,random_state=48)
train_set_z5, test_set_z5 = train_test_split(z_t1[4],test_size=.2,random_state=48)
train_set_z6, test_set_z6 = train_test_split(z_t1[5],test_size=.2,random_state=48)
train_set_z = [train_set_z1,train_set_z2,train_set_z3,train_set_z4,train_set_z5,train_set_z6]
test_set_z = [test_set_z1, test_set_z2,test_set_z3, test_set_z4,test_set_z5, test_set_z6]

#reorganizing the data for ease of axis (not needed could just call the array itself)
Data = {'x': train_set_x,
        'y': train_set_y,
        'z_1': train_set_z1,
        'z_2': train_set_z2,
        'z_3': train_set_z3,
        'z_4': train_set_z4,
        'z_5': train_set_z5,
        'z_6': train_set_z6,
        }
df = pd.DataFrame(Data,columns=['x','y','z_1','z_2','z_3','z_4','z_5','z_6'])
input=df[['x','y']]

#in order to look at more outputs, just add the Z_number value for it to the list bellow

z_variable_selection = 1 # change to mach z_value
output = df[['z_1']]     # change to data set
regression=linear_model.LinearRegression()
regression.fit(input.values,output.values)

#The print statment bellow can be uncommented in order to allow it to print the itercepts and coeficents of the function
#print(regression.intercept_, regression.coef_)

#in here are the input values used to predict the function
start=timeit.default_timer()
prediction_data = list()
for x in range(len(test_set_x)):
    prediction_data.append(float(regression.predict([[test_set_x[x], test_set_y[x]]])))

stop = timeit.default_timer()
print('Time: ', stop - start)

summation = 0
for test_value in range(len(test_set_x)):
    summation = summation + (Data["z_1"][test_value]-prediction_data[test_value])**2 # to match above value
MSE = 1/len(test_set_x)*summation
print(MSE)

fig = plt.figure()
ax_real = fig.add_subplot(1,2,1,projection='3d')
ax_prediction = fig.add_subplot(1,2,2, projection = '3d')
surf_real = ax_real.plot_trisurf(test_set_x,test_set_y,test_set_z[z_variable_selection-1], cmap=cm.coolwarm,linewidth=0, antialiased=False)
surf_prediction= ax_prediction.plot_trisurf(test_set_x, test_set_y, prediction_data, cmap=cm.coolwarm,linewidth=0, antialiased=False)
cbaxes = fig.add_axes([0.1, 0.3, 0.01, 0.4]) 
fig.colorbar(surf_real, shrink=0.5, aspect=5, cax = cbaxes )
ax_real.set_xlabel('Vdc1, volts')
ax_real.set_ylabel('Vdc2, volts')
ax_real.set_zlabel('Angle, degrees')
ax_prediction.set_xlabel('Vdc1, volts')
ax_prediction.set_ylabel('Vdc2, volts')
ax_prediction.set_zlabel('Angle, degrees')
plt.show()