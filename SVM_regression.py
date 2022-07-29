import numpy as np
import numpy as np
from openpyxl import load_workbook
import numpy as np
import timeit
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
from matplotlib import cm
import pandas as pd
from sklearn.metrics import mean_squared_error
from sklearn.svm import LinearSVR
from sklearn.model_selection import cross_val_score

z_variable_selection = 0 # change this variable to corrispond with the theta value form 0-5
Epsilon = 1.5            #Change this variable to corrispond with the epsilon value wanted

wb = load_workbook(r"C:\Users\alqas\Research\SHEtable2Vdc.xlsx")
ws = wb.active

x_vdc1=list()
y_vdc2=list()
z_t1=[[],[],[],[],[],[]]
#collecting the data from the columns we need
for col in ws.iter_cols(min_row=2, max_col=2 ,min_col=2, values_only=True):
    for row in col:
        x_vdc1.append(row)
for col in ws.iter_cols(min_row=2, max_col=3 ,min_col=3, values_only=True):
    for row in col:
        y_vdc2.append(row)
for g in range(1,7):
    for col in ws.iter_cols(min_row=2, max_col=g+3 ,min_col=g+3, values_only=True):
        for row in col:
            z_t1[g-1].append(row)

# resizing the data to fit the model
df = pd.DataFrame({"x_1":np.array(x_vdc1).reshape(496,), "x_2":np.array(y_vdc2).reshape(496,), "y":np.array(z_t1[z_variable_selection]).reshape(496,)}, index=range(0,496))

#sepperating inputs and outputs
X, y = df[["x_1", "x_2"]], df["y"]

#splitting test and training data
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=48)

#fitting LinearSVD regression
LinearSVD_Reg = LinearSVR(epsilon =Epsilon )
LinearSVD_Reg.fit(X_train,y_train)

#predicting values for test set
start=timeit.default_timer()  
LinearSVD_predicted = LinearSVD_Reg.predict(X_test)

#finding Mean squared error
LinearSVD_rmse = np.sqrt(mean_squared_error(y_test, LinearSVD_predicted))
stop = timeit.default_timer()

#reformating input data for graphing
x_test = X_test["x_1"]
Y_test = X_test["x_2"]
differance = LinearSVD_predicted - y_test

#graphing expected vs predicted data
fig = plt.figure()
ax_real = fig.add_subplot(1,3,1,projection='3d')
ax_prediction = fig.add_subplot(1,3,2, projection = '3d')
ax_differance = fig.add_subplot(1,3,3, projection = '3d')
surf_real = ax_real.plot_trisurf(x_test,Y_test,y_test, cmap=cm.coolwarm,linewidth=0, antialiased=False)
surf_prediction= ax_prediction.plot_trisurf(x_test,Y_test, LinearSVD_predicted, cmap=cm.coolwarm,linewidth=0, antialiased=False)
surf_differance = ax_differance.plot_trisurf(x_test,Y_test, differance, cmap=cm.coolwarm,linewidth=0, antialiased=False)
cbaxes = fig.add_axes([0.1, 0.3, 0.01, 0.4]) 
fig.colorbar(surf_real, shrink=0.5, aspect=5, cax = cbaxes )
ax_real.set_xlabel('Vdc1, volts')
ax_real.set_ylabel('Vdc2, volts')
ax_real.set_zlabel('Angle, degrees')
ax_prediction.set_xlabel('Vdc1, volts')
ax_prediction.set_ylabel('Vdc2, volts')
ax_prediction.set_zlabel('Angle, degrees')
ax_differance.set_xlabel('Vdc1, volts')
ax_differance.set_ylabel('Vdc2, volts')
ax_differance.set_zlabel('Angle, degrees')

#cross validation 
scores = cross_val_score(LinearSVD_Reg,X,y,scoring="neg_mean_squared_error",cv=10)
LinearSVD_CV = np.sqrt(-scores) 


## EXECUTABLE ITEMS

#predicted values
print("prediction values",LinearSVD_predicted)
# RMSE of cross validation
print("RMSE for set",LinearSVD_rmse)
# time prediction/execution
print('Time for prediction: ', stop - start)
# Cross validation information
print("the value of cross validation value",LinearSVD_CV.mean())
#show graph
plt.show()